# Francisco, Sahar, Edward
# ReinforceAlgorithm Class: Solver.
import environmentModel as em
import torch
import torch.nn as nn
from torch.distributions import Categorical
import sys
import numpy as np  # numerical python
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
# printoptions: output limited to 2 digits after decimal point
np.set_printoptions(precision=2, suppress=False)


class Solver():

    def __init__(self, numberEpisodes, Model, discountFactor, numberIterations):
        self.numberEpisodes = numberEpisodes
        self.env = Model
        self.gamma = discountFactor
        self.numberIterations = numberIterations
        self.bestPolicy = None

    def runBestPolicy(self):
        """
            Run best policy from the Reinforcement Learning Algorithm. It needs to be used after training.
        """

        state, reward, done = self.env.reset()
        returns = 0
        while not done:
            prev_state = state
            probs = self.bestPolicy(prev_state)
            distAction = Categorical(probs)
            action = distAction.sample()

            state, reward, done = self.env.step(prev_state, action.item())
            returns = returns + reward

        return returns

    def write_to_excel(self, new_row):
        """
        row includes:  # h/l(cost)  ep	adversary	return  advReturn	loss	actor	critic	lr	gamma	hist	clc   actions  nn_name total_stages	num_actions return_against_adversaries
        """

        path = 'results.xlsx'
        wb = load_workbook(path)
        sheet = wb.active
        row = 2
        col = 1
        sheet.insert_rows(idx=row)

        for i in range(len(new_row)):
            sheet.cell(row=row, column=col+i).value = new_row[i]
        wb.save(path)


class ReinforceAlgorithm(Solver):
    """
        Model Solver.
    """

    def __init__(self, Model, neuralNet, numberIterations, numberEpisodes, discountFactor) -> None:
        super().__init__(numberEpisodes, Model, discountFactor, numberIterations)

        self.env.adversaryReturns = np.zeros(numberEpisodes)
        self.neuralNetwork = neuralNet
        self.policy = None
        self.optim = None
        self.bestAverageRetu = 0

        self.returns = []
        self.loss = []
        # self.returns = np.zeros((numberIterations, numberEpisodes))
        # self.loss = np.zeros((numberIterations, numberEpisodes))

    def resetPolicyNet(self):
        """
            Reset Policy Neural Network.
        """
        self.policy, self.optim = self.neuralNetwork.reset()

    def saveResult(self):
        pass

    def solver(self, print_step=10_000, prob_break_limit_ln=-0.001, clc=0.1):
        """
            Method that performs Monte Carlo Policy Gradient algorithm. 
        """

        for iteration in range(self.numberIterations):
            self.resetPolicyNet()

            self.returns.append([])
            self.loss.append([])

            for episode in range(self.numberEpisodes):

                episodeMemory = list()
                state, reward, done = self.env.reset()
                returns = 0
                probs_lst = []
                value_lst = []
                while not done:
                    prevState = state
                    normPrevState = self.normalizeState(prevState)
                    probs, value = self.policy(normPrevState)
                    distAction = Categorical(probs)
                    probs_lst.append(probs)
                    value_lst.append(value)
                    action = distAction.sample()

                    # if episode % 1000 == 0:
                    #     print("-"*30)
                    #     print("probs= ", probs)

                    state, reward, done = self.env.step(
                        prevState, action.item())
                    returns = returns + reward
                    episodeMemory.append((normPrevState, action, reward))

                states = torch.stack([item[0] for item in episodeMemory])
                actions = torch.tensor([item[1] for item in episodeMemory])
                rewards = torch.tensor([item[2] for item in episodeMemory])

                action_probs = torch.stack(probs_lst)
                action_dists = Categorical(action_probs)

                action_logprobs = action_dists.log_prob(actions)

                discReturns = (self.returnsComputation(
                    rewards, episodeMemory))

                discValues = (self.returnsComputation(
                    value_lst, episodeMemory))

                actor_loss = - (torch.sum((discReturns-(discValues).detach())*action_logprobs)) / \
                    len(episodeMemory)
                critic_loss = torch.sum(
                    torch.pow(discValues-discReturns, 2))/len(episodeMemory)

                loss = actor_loss + clc*critic_loss

                shouldBreak = torch.all((action_logprobs >
                                         prob_break_limit_ln))

                if (episode % print_step == 0) or shouldBreak:
                    print("-"*50)
                    print(episode, "  adversary: ", self.env.adversaryMode)
                    print("  actions: ", actions)

                    print("loss= ", loss, "  , actor= ", actor_loss,
                          "  , critic= ", critic_loss, "  , return= ", returns)
                    # print("states= ", states)
                    print("probs of actions: ", torch.exp(action_logprobs))
                    print("discounted returns: ", discReturns)
                    print("discounted values: ", discValues)
                    print("stage values: ", torch.cat(
                        value_lst).detach().numpy())
                    # print("shouldBreak:", shouldBreak.item())
                    # print("actionProbsDist",action_probs)
                    # print("action_dists",action_dists)
                    # print("action_logprobs",action_logprobs)

                self.optim.zero_grad()

                loss.backward()
                self.optim.step()

                # sum of the our player's rewards  rounds 0-25
                # self.returns[iteration][episode] = returns
                # self.loss[iteration][episode] = loss
                self.returns[iteration].append(returns)
                self.loss[iteration].append(loss.item())

                # all probs >0.999 means coverged? break
                if shouldBreak:
                    # self.returns[iteration] = self.returns[iteration][0:episode]
                    # self.loss[iteration] = self.loss[iteration][0:episode]
                    break

            # averageRetu = (
            #     (self.returns[iteration]).sum())/(self.numberEpisodes)
            # if (self.bestPolicy is None) or (averageRetu > self.bestAverageRetu):
            #     self.bestPolicy = self.policy
            #     self.bestAverageRetu = averageRetu

            advModeNames = ""
            for i in range(len(self.env.adversaryProbs)):
                if self.env.adversaryProbs[i] != 0:
                    tmp = "{:.1f}".format(self.env.adversaryProbs[i])
                    advModeNames += f"{(em.AdversaryModes(i)).name}-{tmp}-"

            name = f"ep {len(self.returns[iteration])}, {advModeNames}, {self.env.advHistoryNum} hist, {self.neuralNetwork.lr} lr"
            self.neuralNetwork.save(name=name)

            # h/l(cost)  ep	adversary	return	advReturn  loss	 actor	critic	lr	gamma	hist	clc   actions  nn_name total_stages	num_actions return_against adversaries
            new_row = [('h' if self.env.costs[0] > self.env.costs[1] else 'l'), len(self.returns[iteration]), str(self.env.adversaryProbs), returns, sum(self.env.profit[1]), loss.item(), actor_loss.item(
            ), critic_loss.item(), self.neuralNetwork.lr, self.gamma, self.env.advHistoryNum, clc, str(actions), name, self.env.T, self.neuralNetwork.num_actions]

            for advmode in em.AdversaryModes:
                new_row.append(
                    np.array(self.playTrainedAgent(advmode, 10)).mean())

            self.write_to_excel(new_row)

            plt.scatter(
                range(len(self.returns[iteration])), self.returns[iteration])
            plt.show()

    def returnsComputation(self, rewards, episodeMemory):
        """
        Method computes vector of returns for every stage. The returns are the cumulative rewards from that stage.
        output:tensor

        """

        discRewards = torch.zeros(len(episodeMemory))
        discRewards[-1] = rewards[-1]
        for i in range(len(episodeMemory)-2, -1, -1):
            discRewards[i] = rewards[i] + self.gamma*discRewards[i+1]
        return discRewards

        # return torch.tensor([torch.sum(rewards[i:] * (self.gamma ** torch.arange(0, (len(episodeMemory)-i)))) for i in range(len(episodeMemory))])

    def normalizeState(self, state):
        normalized = [0]*len(state)
        normalized[0] = (state[0]+1)/(self.env.T)
        for i in range(1, len(state)):
            normalized[i] = state[i]/(self.env.totalDemand)
        return torch.tensor(normalized)
        # return state

    def playTrainedAgent(self, advMode, iterNum):
        advProbs = torch.zeros(len(em.AdversaryModes))
        advProbs[int(advMode.value)] = 1
        game = em.Model(totalDemand=self.env.totalDemand,
                        tupleCosts=self.env.costs,
                        totalStages=self.env.T, advHistoryNum=self.env.advHistoryNum, adversaryProbs=advProbs)
        returns = np.zeros(iterNum)

        # self.policy, _ = self.neuralNetwork.reset()
        # self.neuralNetwork.load(NN_name)

        for episode in range(iterNum):

            state, reward, done = game.reset()
            retu = 0

            while not done:
                prevState = state
                normPrevState = self.normalizeState(prevState)
                probs, _ = self.policy(normPrevState)
                distAction = Categorical(probs)

                action = distAction.sample()

                # if episode % 1000 == 0:
                #     print("-"*30)
                #     print("probs= ", probs)

                state, reward, done = game.step(
                    prevState, action.item())
                retu += reward

                # prevState = state
                # normPrevState = self.normalizeState(prevState)
                # probs, _ = self.policy(normPrevState)
                # distAction = Categorical(probs)
                # action = distAction.sample()

                # state, reward, done = game.step(
                #     prevState, action.item())
                # retu = retu + reward
                # episodeMemory.append((normPrevState, action, reward))

            # states = torch.stack([item[0] for item in episodeMemory])
            # actions = torch.tensor([item[1] for item in episodeMemory])
            # rewards = torch.tensor([item[2] for item in episodeMemory])

            # print(f"iteration {episode} return= {retu} \n\t actions: {actions}")

            # sum of the our player's rewards  rounds 0-25
            returns[episode] = retu

            print('returns', retu)

            print("profits sum: ", sum(game.profit[0]), sum(game.profit[1]))

        # plt.plot(returns)
        # plt.show()

        return returns
