# Francisco, Sahar, Edward
# ReinforceAlgorithm Class: Solver.
import environmentModelBase as model
import torch
import torch.nn as nn
from torch.distributions import Categorical
import sys
import numpy as np  # numerical python
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
import time
# printoptions: output limited to 2 digits after decimal point
np.set_printoptions(precision=2, suppress=False)


class Solver():

    def __init__(self, numberEpisodes, Model, discountFactor, numberIterations):
        self.numberEpisodes = numberEpisodes
        self.env = Model
        self.gamma = discountFactor
        self.numberIterations = numberIterations
        self.bestPolicy = None
        self.probBreakLn = -0.001
        


    def write_to_excel(self, new_row):
        """
        row includes:  # ep	adversary	return	advReturn loss	actor	critic	lr	gamma	hist	clc   actions probs  nn_name  total_stages	action_step num_actions return_against_adversaries
        """

        path = 'results.xlsx'
        wb = load_workbook(path)
        sheet = wb.active
        row = 2
        col = 1
        sheet.insert_rows(idx=row)

        for i in range(len(new_row)):
            sheet.cell(row=row, column=col+i).value = new_row[i]
        wb.save(path)
        # print(new_row)


class ReinforceAlgorithm(Solver):
    """
        Model Solver.
    """

    def __init__(self, Model, neuralNet, numberIterations, numberEpisodes, discountFactor,actionStep=1) -> None:
        super().__init__(numberEpisodes, Model, discountFactor, numberIterations)

        self.env.adversaryReturns = np.zeros(numberEpisodes)
        self.neuralNetwork = neuralNet
        self.policy = None
        self.optim = None
        self.bestAverageRetu = 0

        self.returns = []
        self.loss = []
        self.actionStep, self.neuralNetwork.action_step=actionStep, actionStep
        # self.returns = np.zeros((numberIterations, numberEpisodes))
        # self.loss = np.zeros((numberIterations, numberEpisodes))

        self.dataRow=[]

    def resetPolicyNet(self):
        """
            Reset Policy Neural Network.
        """
        self.policy, self.optim = self.neuralNetwork.reset()
    def loadPolicyNet(self, name):
        """
            Load Policy Neural Network.
        """
        self.resetPolicyNet()
        self.neuralNetwork.load(name)


    # def saveResult(self):
    #     pass

    def returnsComputation(self, rewards, episodeMemory=None):
        """
        Method computes vector of returns for every stage. The returns are the cumulative rewards from that stage.
        output:tensor

        """

        discRewards = torch.zeros(len(rewards))
        discRewards[-1] = rewards[-1]
        for i in range(len(rewards)-2, -1, -1):
            discRewards[i] = rewards[i] + self.gamma*discRewards[i+1]
        return discRewards

        # return torch.tensor([torch.sum(rewards[i:] * (self.gamma ** torch.arange(0, (len(episodeMemory)-i)))) for i in range(len(episodeMemory))])

    

    def solver(self, print_step=50_000, options=[1, 10000, 1, 1], converge_break=False):
        """
        print_step=None means no printing
        """
        self.returns = []
        self.loss = []

        # fig, axs = plt.subplots(max(self.numberIterations,2), 2, figsize=(15, 6*self.numberIterations))

        for iter in range(self.numberIterations):

            self.resetPolicyNet()
            self.returns.append([])
            self.loss.append([])

            for stage in range(self.env.T-1, -1, -1):
                self.learn_stage_onwards(iter,stage=stage, episodes=int(self.numberEpisodes*(self.env.T-stage)), print_step=print_step, options=options,
                                prob_break_limit_ln=(self.probBreakLn if converge_break else None), write_save=True if stage==0 else False)
                
            # axs[iter][0].scatter(range(len(self.returns[iter])), self.returns[iter])
            # axs[iter][1].scatter(range(len(self.loss[iter])), self.loss[iter])

        # plt.show()
            

   

    def learn_stage_onwards(self,iter,stage, episodes, print_step=10_000, prob_break_limit_ln=None, options=[1, 10000, 1, 2], lr=None, just_stage=False, write_save=False):
        """
            Method that just learns the actions of stages after the input stage. 

        """
        if self.policy == None:
            self.resetPolicyNet()

        if lr is not None:
            self.optim.lr = lr


        for episode in range(episodes):

            episodeMemory = list()
            state, reward, done = self.env.reset()
            returns = 0
            probs_lst = []
            while not done:
                prevState = state
                normPrevState = self.env.normalizeState(prevState)
                probs = self.policy(normPrevState)
                distAction = Categorical(probs)
                probs_lst.append(probs)

                action = distAction.sample()

                # if episode % 1000 == 0:
                # print("-"*60)
                # print("probs= ", probs)

                state, reward, done = self.env.step(
                    self.env.compute_price(action=action.item(),actionStep=self.actionStep))
                returns = returns + reward
                episodeMemory.append((normPrevState, action, reward))

            if episode == 0:
                probs_lst_pre = probs_lst

            states = torch.stack([item[0] for item in episodeMemory])
            actions = torch.tensor([item[1] for item in episodeMemory])
            rewards = torch.tensor([item[2]
                                   for item in episodeMemory])/options[1]

            action_probs = torch.stack(probs_lst)
            action_dists = Categorical(action_probs)

            action_logprobs = action_dists.log_prob(actions)

            

            actionsLogProbs = action_logprobs[stage:]
            discRewards = self.returnsComputation(rewards=rewards)
            baseRewards = self.computeBase(
                self.env.prices[1], initDemand=self.env.demandPotential[0][stage], startStage=stage)/options[1]
            baseDiscReturns = discRewards-baseRewards
            finalReturns = baseDiscReturns[stage:]

            if just_stage:
                loss = -baseDiscReturns[stage]*action_logprobs[stage]
            else:
                loss = -(torch.sum(finalReturns*actionsLogProbs))

            
            shouldBreak = False

            if prob_break_limit_ln is not None and torch.all(action_logprobs[stage:] > prob_break_limit_ln):
                shouldBreak = True

            if (print_step is not None) and ((episode % print_step == (print_step-1)) or shouldBreak):
                print("-"*20)
                

                print("iter ",iter," stage ",stage," ep ",episode, "  adversary: ", self.env.adversary)
                print("  actions: ", actions* self.actionStep)

                print("loss= ", loss,"return= ", returns)
                # print("states= ", states)
                print("probs of actions: ", torch.exp(
                    action_logprobs))
                # print("action_logprobs: ", action_logprobs)
                # print("probs=", probs_lst)
                # print("discounted returns: ", baseDiscReturns)
                # print("rewards: ", rewards)
                # print("finalReturns: ", finalReturns)

                # print("nn 1st layer",self.policy[0].weight)
                # print("nn 2nd layer",self.policy[2].weight)
                # print("shouldBreak:", shouldBreak.item())
                # print("actionProbsDist",action_probs)
                # print("action_dists",action_dists)
                # print("action_logprobs",action_logprobs)

            probs_lst_pre = probs_lst

            self.policy.zero_grad()
            loss.backward()
            self.optim.step()

            # if episode != 0:
            #     self.meanStageValue = (
            #         (self.meanStageValue*episode)+rewards)/(episode+1)

            # sum of the our player's rewards  rounds 0-25
            # self.returns[iteration][episode] = returns
            # self.loss[iteration][episode] = loss
            self.returns[iter].append(returns)
            self.loss[iter].append(loss.item())

            # all probs >0.999 means coverged? break
            if shouldBreak:
                # self.returns[iteration] = self.returns[iteration][0:episode]
                # self.loss[iteration] = self.loss[iteration][0:episode]
                break
        
           
        name = f"{int(time.time())}"
            # ep	adversary	return  advReturn	loss  lr	gamma	hist  actions   probs  nn_name  total_stages	action_step  num_actions   return_against_adversaries
        self.dataRow= [len(self.returns[iter]), str(self.env.advMixedStrategy), returns, sum(self.env.profit[1]), loss.item(), self.neuralNetwork.lr, self.gamma, self.env.stateAdvHistory, str(actions*self.actionStep), str((torch.exp(action_logprobs)).detach().numpy()), name, self.env.T, self.neuralNetwork.action_step, self.neuralNetwork.num_actions]

        self.neuralNetwork.nn_name=name
            # for advmode in model.AdversaryModes:
            #     new_row.append(np.array((self.playTrainedAgent(advmode,10))[0]).mean())

            
            
    def write_nn_data(self):
        """
        writes the data in excel and saves nn
        """
        

            # self.name = f"{[self.neuralNetwork.lr, self.gamma,clc]}-stage {stage}-{int(time.time())}"
        self.neuralNetwork.save()
        print(self.neuralNetwork.nn_name, "saved")
        # ep	adversary	return  advReturn	loss  lr	gamma	hist  actions   probs  nn_name  total_stages	action_step  num_actions   return_against_adversaries
        
        self.write_to_excel(self.dataRow)
        
        

    def computeBase(self, advPrices, startStage=0, initDemand=None):
        """
        discounted rewards when we play myopic against the adversary

        """
        if initDemand is None:
            initDemand = (self.env.getState(stage=0))[self.env.T]
        profit = torch.zeros(self.env.T)
        demand = initDemand
        for i in range(startStage, self.env.T):
            price = (demand + self.env.costs[0])/2

            profit[i] = (demand-price)*(price-self.env.costs[0])
            demand += (advPrices[i]-price)/2
        return self.returnsComputation(rewards=profit)

    # def myopic_price(demand,cost):
    #     return (demand + cost)/2

    def playTrainedAgent(self, adversary, iterNum):
        """
        Current trained NN will plays against the adversary's strategy, without learning.
        """

        game = model.Model(totalDemand=self.env.totalDemand,
                           tupleCosts=self.env.costs,
                           totalStages=self.env.T, advMixedStrategy=adversary ,stateAdvHistory=self.neuralNetwork.adv_hist)
        returns = torch.zeros(2,iterNum)
        for episode in range(iterNum):

            # episodeMemory = list()
            state, reward, done = game.reset()
            retu = 0

            while not done:
                prevState = state
                normPrevState = self.env.normalizeState(prevState)
                probs= self.policy(normPrevState)
                distAction = Categorical(probs)
                action = distAction.sample()

                state, reward, done = game.step(
                 game.compute_price(action=action.item(),actionStep=self.actionStep))
                retu = retu + reward
                # episodeMemory.append((normPrevState, action, reward))

            # states = torch.stack([item[0] for item in episodeMemory])
            # actions = torch.tensor([item[1] for item in episodeMemory])
            # rewards = torch.tensor([item[2] for item in episodeMemory])

            # print(f"iteration {episode} return= {retu} \n\t actions: {actions}")

            # sum of the our player's rewards  rounds 0-25
            returns[0][episode], returns[1][episode] = sum(game.profit[0]), sum(game.profit[1])
        
        return returns

        # plt.plot(returns)
        # plt.show()
