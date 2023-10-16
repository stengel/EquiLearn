
import environmentModelBase as model
import torch
import torch.nn as nn
from torch.distributions import Categorical
from enum import Enum
import numpy as np  # numerical python
from openpyxl import load_workbook
import time
import globals as gl
from collections import deque, namedtuple
import random
from multiprocessing import Pool

# printoptions: output limited to 2 digits after decimal point
np.set_printoptions(precision=2, suppress=False)


class Solver():

    def __init__(self, numberEpisodes, Model, discountFactor, numberIterations):
        self.numberEpisodes = numberEpisodes
        self.env = Model
        self.gamma = discountFactor
        self.numberIterations = numberIterations
        self.bestPolicy = None
        self.probBreakLn = -0.001
        self.costs = self.env.costs

    def write_to_excel(self, new_row):
        """
        row includes:  # ep	adversary	return  advReturn	loss  lr	gamma	hist  actions   probs  nn_name  total_stages	action_step  num_actions   agent's_prices
        # adv's_prices  agent's_demands   adv's_demands   return_against_adversaries
        """

        path = 'results.xlsx'
        wb = load_workbook(path)
        sheet = wb.active
        row = 2
        col = 1
        sheet.insert_rows(idx=row)

        for i in range(len(new_row)):
            sheet.cell(row=row, column=col+i).value = new_row[i]
        wb.save(path)
        # print(new_row)


class ReinforceAlgorithm(Solver):
    """
        Model Solver.
    """

    def __init__(self, Model, neuralNet, numberIterations, numberEpisodes, discountFactor) -> None:
        super().__init__(numberEpisodes, Model, discountFactor, numberIterations)

        self.env.adversaryReturns = np.zeros(numberEpisodes)
        self.neuralNetwork = neuralNet
        self.policy = None
        self.optim = None

        self.returns = []
        self.loss = []

        self.dataRow = []

    def reset_policy_net(self):
        """
            Reset Policy Neural Network.
        """
        self.policy, self.optim = self.neuralNetwork.reset()

    def load_policy_net(self, name):
        """
            Load Policy Neural Network.
        """
        self.reset_policy_net()
        self.neuralNetwork.load(name)

    # def saveResult(self):
    #     pass

        # return torch.tensor([torch.sum(rewards[i:] * (self.gamma ** torch.arange(0, (len(episodeMemory)-i)))) for i in range(len(episodeMemory))])

    def solver(self):
        """
        print_step=None means no printing
        """
        self.returns = []
        self.loss = []

        # fig, axs = plt.subplots(max(self.numberIterations,2), 2, figsize=(15, 6*self.numberIterations))

        for iter in range(self.numberIterations):
            replay_buffer = ReplayBuffer(gl.replay_buffer_size)
            self.reset_policy_net()
            self.returns.append([])
            self.loss.append([])

            for stage in range(self.env.T-1, -1, -1):
                episodes = self.compute_num_episodes(stage)
                if stage < self.env.T-1:
                    self.play_from_buffer(stage=stage, replay_buffer=replay_buffer, episodes=int(
                        gl.buffer_play_coefficient*episodes))

                new_espisodes = int((1-gl.buffer_play_coefficient) *
                                    episodes) if (stage < gl.total_stages-1) else episodes

                self.learn_stage_onwards(
                    iter=iter, stage=stage, episodes=new_espisodes, replay_buffer=replay_buffer)

            # axs[iter][0].scatter(range(len(self.returns[iter])), self.returns[iter])
            # axs[iter][1].scatter(range(len(self.loss[iter])), self.loss[iter])

        # plt.show()

    def compute_num_episodes(self, stage):
        return int(self.numberEpisodes*((gl.total_stages-stage)**1.2))
        # return int(self.numberEpisodes*((gl.total_stages-stage)))

    def learn_stage_onwards(self, iter, stage, episodes, replay_buffer):
        """
            Method that just learns the actions of stages after the input stage. 

        """
        # if stage<gl.total_stages-1:
        #     buffer_loss = play_from_buffer(stage=stage, replay_buffer=replay_buffer, episodes=int(
        #         gl.buffer_play_coefficient*episodes))
        #     for loss in buffer_loss:
        #         self.policy.zero_grad()
        #         loss.backward()
        #         self.optim.step()

        for episode in range(episodes):

            episode_memory = list()
            state, reward, done = self.env.reset()
            returns = 0
            probs_lst = []
            while not done:
                prev_state = state
                norm_prev_state = normalize_state(prev_state)
                probs = self.policy(norm_prev_state)
                dist_action = Categorical(probs)
                probs_lst.append(probs)

                action = dist_action.sample()

                # if episode % 1000 == 0:
                # print("-"*60)
                # print("probs= ", probs)

                state, reward, done = self.env.step(
                    compute_price(self.env.demandPotential[0][self.env.stage], self.env.costs[0], action=action.item(), action_step=gl.action_step))
                returns = returns + reward
                episode_memory.append((norm_prev_state, action, reward))

            # states = torch.stack([item[0] for item in episode_memory])
            actions = torch.tensor([item[1] for item in episode_memory])
            rewards = torch.tensor([item[2]
                                   for item in episode_memory])/gl.rewards_division_const

            action_probs = torch.stack(probs_lst)
            action_dists = Categorical(action_probs)

            action_log_probs = action_dists.log_prob(actions)
            replay_buffer.add_game(
                actions, self.env.demandPotential[0], self.env.prices[0], self.env.prices[1],  rewards)

            action_log_probs_cut = action_log_probs[stage:]
            base_disc_returns = returns_computation(rewards=rewards) - (compute_base(
                agent_cost=self.costs[0], adv_prices=self.env.prices[1], stage_demand=self.env.demandPotential[0][stage], start_stage=stage)/gl.rewards_division_const)
            final_returns = base_disc_returns[stage:]

            # if just_stage:
            #     loss = -baseDiscReturns[stage]*action_logprobs[stage]
            # else:
            loss = loss_function(final_returns, action_log_probs_cut)

            should_break = False

            if gl.prob_break_limit_ln is not None and torch.all(action_log_probs[stage:] > gl.prob_break_limit_ln):
                should_break = True

            if (gl.print_step is not None) and ((episode % gl.print_step == (gl.print_step-1)) or should_break):
                print("-"*20)

                print("iter ", iter, " stage ", stage, " ep ",
                      episode, "  adversary: ", self.env.adversary)
                print("  actions: ", actions * self.actionStep)

                print("loss= ", loss, "return= ", returns)
                # print("states= ", states)
                print("probs of actions: ", torch.exp(
                    action_log_probs))
                # print("action_logprobs: ", action_logprobs)
                # print("probs=", probs_lst)
                # print("discounted returns: ", baseDiscReturns)
                # print("rewards: ", rewards)
                # print("finalReturns: ", finalReturns)

                # print("nn 1st layer",self.policy[0].weight)
                # print("nn 2nd layer",self.policy[2].weight)
                # print("shouldBreak:", shouldBreak.item())
                # print("actionProbsDist",action_probs)
                # print("action_dists",action_dists)
                # print("action_logprobs",action_logprobs)

            self.policy.zero_grad()
            loss.backward()
            self.optim.step()

            # if episode != 0:
            #     self.meanStageValue = (
            #         (self.meanStageValue*episode)+rewards)/(episode+1)

            # sum of the our player's rewards  rounds 0-25
            # self.returns[iteration][episode] = returns
            # self.loss[iteration][episode] = loss
            self.returns[iter].append(returns)
            self.loss[iter].append(loss.item())

            # all probs >0.999 means coverged? break
            if should_break:
                # self.returns[iteration] = self.returns[iteration][0:episode]
                # self.loss[iteration] = self.loss[iteration][0:episode]
                break

        name = f"{int(time.time())}"
        # ep  costs adversary	return  advReturn	loss  lr	gamma	hist  actions   probs  nn_name  total_stages	action_step  num_actions   agent's_prices
        # adv's_prices  agent's_demands   adv's_demands   return_against_adversaries
        self.dataRow = [len(self.returns[iter]), str(self.costs),str(self.env.advMixedStrategy), returns, sum(self.env.profit[1]), loss.item(), self.neuralNetwork.lr, self.gamma, self.env.stateAdvHistory, str(
            actions*gl.action_step), str((torch.exp(action_log_probs)).detach().numpy()), name, self.env.T, gl.action_step, self.neuralNetwork.num_actions, str(self.env.prices[0]), str(self.env.prices[1]), str(self.env.demandPotential[0]), str(self.env.demandPotential[1])]

        self.neuralNetwork.name = name
        # for advmode in model.AdversaryModes:
        #     new_row.append(np.array((self.playTrainedAgent(advmode,10))[0]).mean())

    def play_from_buffer(self, stage, replay_buffer, episodes):

        samples = replay_buffer.sample_game(episodes)
        # # field_names==["0 actions", "1 agent_demands", "2 agent_prices","3 adv_prices", "4 rewards"]
        for sample in samples:
            action_log_probs_cut = []
            for s in range(stage, gl.total_stages):
                state = model.define_state(s, gl.total_demand, gl.total_stages,
                                           self.costs[0], sample[2], sample[3], sample[1], gl.num_adv_history)
                norm_state = normalize_state(state)
                probs = self.policy(norm_state)
                dist_action = Categorical(probs)
                action_log_probs_cut.append(
                    dist_action.log_prob((sample[0])[s]))

            base_disc_returns = returns_computation(rewards=sample[4]) - (compute_base(agent_cost=self.costs[0],
                                                                                       stage_demand=(sample[1])[stage], adv_prices=sample[3], start_stage=stage)/gl.rewards_division_const)

            loss = loss_function(
                base_disc_returns[stage:], torch.stack(action_log_probs_cut))
            self.policy.zero_grad()
            loss.backward()
            self.optim.step()

    def write_nn_data(self, prefix=""):
        """
        writes the data in excel and saves nn
        """
        self.neuralNetwork.name=f"{prefix}_{self.neuralNetwork.name}"
        # self.name = f"{[self.neuralNetwork.lr, self.gamma,clc]}-stage {stage}-{int(time.time())}"
        self.neuralNetwork.save()
        # print(self.neuralNetwork.name, "saved")
       # ep	 costs adversary	return  advReturn	loss  lr	gamma	hist  actions   probs  nn_name  total_stages	action_step  num_actions   agent's_prices
        # adv's_prices  agent's_demands   adv's_demands   return_against_adversaries
        self.write_to_excel(self.dataRow)

    # def myopic_price(demand,cost):
    #     return (demand + cost)/2

    def play_trained_agent(self, adversary, iterNum):
        """
        Current trained NN will plays against the adversary's strategy, without learning.
        """

        game = model.Model(totalDemand=self.env.totalDemand,
                           tupleCosts=self.env.costs,
                           totalStages=self.env.T, advMixedStrategy=adversary, stateAdvHistory=gl.num_adv_history)
        returns = torch.zeros(2, iterNum)
        for episode in range(iterNum):

            # episodeMemory = list()
            state, reward, done = game.reset()
            retu = 0

            while not done:
                prevState = state
                normPrevState = normalize_state(prevState)
                probs = self.policy(normPrevState)
                distAction = Categorical(probs)
                action = distAction.sample()

                state, reward, done = game.step(
                    compute_price(game.demandPotential[0][game.stage], game.costs[0], action=action.item(), action_step=gl.action_step))
                retu = retu + reward
                # episodeMemory.append((normPrevState, action, reward))

            # states = torch.stack([item[0] for item in episodeMemory])
            # actions = torch.tensor([item[1] for item in episodeMemory])
            # rewards = torch.tensor([item[2] for item in episodeMemory])

            # print(f"iteration {episode} return= {retu} \n\t actions: {actions}")

            # sum of the our player's rewards  rounds 0-25
            returns[0][episode], returns[1][episode] = sum(
                game.profit[0]), sum(game.profit[1])

        return returns

        # plt.plot(returns)
        # plt.show()


class Strategy():
    """
    strategies can be static or they can come from neural nets. If NN, policy is nn.policy o.w. the static function
    """
    type = None
    env = None
    name = None
    nn = None
    nn_hist = None
    policy = None

    def __init__(self, strategyType, NNorFunc, name, firstPrice=132) -> None:
        """
        Based on the type of strategy, the neuralnet or the Strategy Function  should be given as input. FirstPrice just applies to static strategies
        """
        self.type = strategyType
        self.name = name
        # self._env = environment

        if strategyType == StrategyType.neural_net:
            self.nn = NNorFunc
            self.policy = NNorFunc.policy
            self.nn_hist = gl.num_adv_history
        else:
            self.policy = NNorFunc
            self.firstPrice = firstPrice

    def reset(self):
        pass

    def play(self, environment, player=0):
        """
            Computes the action to be played in the environment, nn.step_action is the step size for pricing less than myopic
        """
        self.env = environment
        if self.type == StrategyType.neural_net:
            state = self.env.get_state(
                self.env.stage, player, adv_hist=gl.num_adv_history)
            normState = normalize_state(state=state)
            probs = self.policy(normState)
            distAction = Categorical(probs)
            action = distAction.sample()
            return compute_price(action=action.item(), action_step=gl.action_step, demand=self.env.demandPotential[player][self.env.stage], cost=self.env.costs[player])

        else:
            return self.policy(self.env, player, self._firstPrice)

    def play_against(self, env, adversary):
        """ 
        self is player 0 and adversary is layer 1. The environment should be specified. action_step for the neural netwroks should be set.
        output: tuple (payoff of low cost, payoff of high cost)
        """
        self.env = env

        state, reward, done = env.reset()
        while env.stage < (env.T):
            prices = [0, 0]
            prices[0], prices[1] = self.play(env, 0), adversary.play(env, 1)
            env.updatePricesProfitDemand(prices)
            env.stage += 1

        return [sum(env.profit[0]), sum(env.profit[1])]

    def to_mixed_strategy(self):
        """
        Returns a MixedStrategy, Pr(self)=1
        """
        mix = MixedStrategy(probablitiesArray=torch.ones(1),
                            strategiesList=[self])

        return mix


class MixedStrategy():
    _strategies = []
    _strategyProbs = None

    def __init__(self, strategiesList=[], probablitiesArray=None) -> None:
        self._strategies = strategiesList
        self._strategyProbs = probablitiesArray

    def set_adversary_strategy(self):
        if len(self._strategies) > 0:
            # adversaryDist = Categorical(torch.tensor(self._strategyProbs))
            if not torch.is_tensor(self._strategyProbs):
                self._strategyProbs = torch.tensor(self._strategyProbs)
            adversaryDist = Categorical(self._strategyProbs)
            strategyInd = (adversaryDist.sample()).item()
            return self._strategies[strategyInd]
        else:
            print("adversary's strategy can not be set!")
            return None

    def __str__(self) -> str:
        s = ""
        for i in range(len(self._strategies)):
            if self._strategyProbs[i] > 0:
                s += f"{self._strategies[i].name}-{self._strategyProbs[i]:.2f},"
        return s


class StrategyType(Enum):
    static = 0
    neural_net = 1


class ReplayBuffer():
    """
        the information of one round of game that is needed for creating the states later, update the nn and compute_base will be saved in buffer. The len of each input array is total_stages
        """

    def __init__(self, max_len):

        self.deque = deque([], maxlen=max_len)
        # self.entry = namedtuple("Entry", field_names=["0 actions", "1 agent_demands", "2 agent_prices","3 adv_prices", "4 rewards"])

    def add_game(self, actions, agent_demands, agent_prices, adv_prices,  rewards):
        # self.entry = namedtuple("Entry", field_names=["0 actions", "1 agent_demands", "2 agent_prices","3 adv_prices", "4 rewards"])
        entry = (actions, agent_demands, agent_prices, adv_prices, rewards)
        self.deque.append(entry)

    def sample_game(self, sample_size):
        """ return samples of tuples=(["0 actions", "1 agent_demands", "2 agent_prices","3 adv_prices", "4 rewards"])"""

        return random.sample(self.deque, min(sample_size, len(self.deque)))


def normalize_state(state):
    # [stage one-hot encoded, agent's demand potential, agent's last price, history of adversary's prices]

    normalized = [0]*len(state)
    for i in range(gl.total_stages):
        normalized[i] = state[i]
    for i in range(gl.total_stages, len(state)):
        normalized[i] = state[i]/(gl.total_demand)
    return torch.tensor(normalized)


def compute_price(demand, cost, action, action_step):
    return model.monopolyPrice(demand, cost) - (action_step * action)


def returns_computation(rewards):
    """
    Method computes vector of returns for every stage. The returns are the cumulative rewards with discount factor from that stage onwards.
    output:tensor

    """
    discRewards = torch.zeros(len(rewards))
    discRewards[-1] = rewards[-1]
    for i in range(len(rewards)-2, -1, -1):
        discRewards[i] = rewards[i] + gl.gamma*discRewards[i+1]
    return discRewards


def compute_base(agent_cost, adv_prices, start_stage=0, stage_demand=None):
    """
    discounted rewards when we play myopic against the adversary from start_stage onwards

    """
    if stage_demand is None:
        stage_demand = gl.total_demand/2
    profit = torch.zeros(gl.total_stages)
    demand = stage_demand
    for i in range(start_stage, gl.total_stages):
        price = (demand + agent_cost)/2
        a = (demand-price)*(price-agent_cost)
        profit[i] = (demand-price)*(price-agent_cost)
        demand += (adv_prices[i]-price)/2
    return returns_computation(rewards=profit)


# def play_from_buffer(stage, replay_buffer, episodes):

#     with Pool() as pool:
#         samples = [replay_buffer.sample_game(
#             int(episodes/gl.num_cores)) for _ in range(gl.num_cores)]
#         stage_repeat = [stage]*gl.num_cores
#         results = pool.starmap(compute_buffer_update,
#                                zip(stage_repeat, samples))
#     return results


# def compute_buffer_update(stage, samples):
    # # field_names=["action_log_probs 0","agent_cost 1", "agent_demands 2", "adv_prices 3", "rewards 4"]
    # loss = []
    # for sample in samples:
    #     action_log_probs_cut = (sample[0])[stage:]
    #     base_disc_returns = returns_computation(rewards=sample[4]) - (compute_base(agent_cost=sample[1],
    #                                                                                     stage_demand=(sample[2])[stage], adv_prices=samples[3], start_stage=stage)/gl.rewardsDivisionConst)
    #     final_returns = base_disc_returns[stage:]

    #     loss.append(-(torch.sum(final_returns*action_log_probs_cut)))

    # return loss

def loss_function(returns, log_probs):

    return -(torch.sum(returns*log_probs))
