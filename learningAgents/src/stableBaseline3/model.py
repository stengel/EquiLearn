from environment import PricingGame
import globals as gl
import classes as cl
import gym
import numpy as np
import os
import time
from openpyxl import load_workbook


def train_agent(costs, adv_mixed_strategy, model, model_name, timesteps, num_timesteps, lr=None):

    iter_name = f"{model_name}-{str(int(time.time()))}"
    models_dir = os.path.join("models", iter_name)
    log_dir = os.path.join("logs", iter_name)

    if not os.path.exists(models_dir):
        os.makedirs(models_dir)

    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    

    env = PricingGame(tuple_costs=costs, adversary_mixed_strategy=adv_mixed_strategy)
    env.reset()

    lr_=(gl.LR if (lr is None) else lr)

    model_ = model('MlpPolicy', env,learning_rate=lr_,verbose=0, tensorboard_log=log_dir, gamma=gl.GAMMA)

    for i in range(num_timesteps):
        model_.learn(total_timesteps=timesteps,
                     reset_num_timesteps=False, tb_log_name=iter_name)
        model_.save(os.path.join(models_dir, str(timesteps*i)))

    # test and write results
    for iter in range(gl.NUM_STOCHASTIC_ITER):
        obs = env.reset()
        done = False

        actions = []
        while not done:
            action, _states = model_.predict(obs)
            obs, reward, done, info = env.step(action)

            actions.append(int(action))
        #   name	ep	costs	adversary	agent_return	adv_return	agent_rewards	actions	agent_prices	adv_prices	agent_demands	adv_demands	    lr	hist	 total_stages	action_step	num_actions  gamma"
        data=[iter_name, timesteps*num_timesteps,("L" if (costs[0]<costs[1]) else "H"), env.adversary_strategy.name, sum(env.profit[0]), sum(env.profit[1]),            str(env.profit[0]), str(actions), str(env.prices[0]), str(env.prices[1]), str(env.demand_potential[0]),str(env.demand_potential[1]), lr_, gl.NUM_ADV_HISTORY, gl.TOTAL_STAGES, gl.ACTION_STEP, gl.NUM_ACTIONS, gl.GAMMA]
        write_to_excel(data)
        

def write_to_excel(new_row):
    """
    row includes:  name	ep	costs	adversary	agent_return	adv_return	agent_rewards	actions	agent_prices	adv_prices	agent_demands	adv_demands	    lr	hist	total_stages	action_step	num_actions  gamma"
    """

    path = 'results.xlsx'
    wb = load_workbook(path)
    sheet = wb.active
    row = 2
    col = 1
    sheet.insert_rows(idx=row)

    for i in range(len(new_row)):
        sheet.cell(row=row, column=col+i).value = new_row[i]
    wb.save(path)
