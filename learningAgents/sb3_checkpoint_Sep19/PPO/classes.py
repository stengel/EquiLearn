from enum import Enum
import numpy as np
import globals as gl
# import torch
# from torch.distributions import Categorical
from openpyxl import load_workbook
from fractions import Fraction
import bimatrix
import time
import os


class BimatrixGame():
    """
    strategies play against each other and fill the matrix of payoff, then the equilibria would be computed using Lemke algorithm
    """

    low_strategies = []
    high_strategies = []
    matrix_A = None
    matrix_B = None

    def __init__(self, low_cost_strategies, high_cost_strategies, env_class) -> None:
        # globals.initialize()
        self.low_strategies = low_cost_strategies
        self.high_strategies = high_cost_strategies
        self.env_class = env_class

    def reset_matrix(self):
        self.matrix_A = np.zeros(
            (len(self.low_strategies), len(self.high_strategies)))
        self.matrix_B = np.zeros(
            (len(self.low_strategies), len(self.high_strategies)))

    def fill_matrix(self):
        self.reset_matrix()

        for low in range(len(self.low_strategies)):
            for high in range(len(self.high_strategies)):
                self.update_matrix_entry(low, high)

    def update_matrix_entry(self, low_index, high_index):
        stratL = self.low_strategies[low_index]
        stratH = self.high_strategies[high_index]
        stratL.reset()
        stratH.reset()

        env = self.env_class(tuple_costs=(
            gl.LOW_COST, gl.HIGH_COST), adversary_mixed_strategy=stratH.to_mixed_strategy())
        payoffs = [stratL.play_against(env, stratH)
                   for _ in range(gl.NUM_STOCHASTIC_ITER)]

        mean_payoffs = (np.mean(np.array(payoffs), axis=0))

        self.matrix_A[low_index][high_index], self.matrix_B[low_index][high_index] = mean_payoffs[0], mean_payoffs[1]

    def write_all_matrix(self):
        # print("A: \n", self._matrix_A)
        # print("B: \n", self._matrix_B)

        output = f"{len(self.matrix_A)} {len(self.matrix_A[0])}\n\n"

        for matrix in [self.matrix_A, self.matrix_B]:
            for i in range(len(self.matrix_A)):
                for j in range(len(self.matrix_A[0])):
                    output += f"{matrix[i][j]:7.0f} "
                output += "\n"
            output += "\n"

        with open("game.txt", "w") as out:
            out.write(output)

        output+="\nlow-cost strategies: \n"
        for strt in self.low_strategies:
            output+=f" {strt.name} "
        output+="\nhigh-cost strategies: \n"
        for strt in self.high_strategies:
            output+=f" {strt.name} "

        with open(f"games/game{int(time.time())}.txt", "w") as out:
            out.write(output)
        

    def add_low_cost_row(self, row_A, row_B):
        self.matrix_A = np.append(self.matrix_A, [row_A], axis=0)
        self.matrix_B = np.append(self.matrix_B, [row_B], axis=0)

    def add_high_cost_col(self, colA, colB):
        self.matrix_A = np.hstack((self.matrix_A, np.atleast_2d(colA).T))
        self.matrix_B = np.hstack((self.matrix_B, np.atleast_2d(colB).T))
        # for j in range(len(self._matrix_A)):
        #     self._matrix_A[j].append(colA[j])
        #     self._matrix_B[j].append(colB[j])

    def compute_equilibria(self):
        self.write_all_matrix()
        game = bimatrix.bimatrix("game.txt")
        equilibria_traces = game.tracing(100, gl.NUM_TRACE_EQUILIBRIA)
        equilibria = []
        for equilibrium in equilibria_traces:
            low_cost_probs, high_cost_probs, low_cost_support, high_cost_support = recover_probs(
                equilibrium)
            low_cost_probabilities = return_distribution(
                len(self.low_strategies), low_cost_probs, low_cost_support)
            high_cost_probabilities = return_distribution(
                len(self.high_strategies), high_cost_probs, high_cost_support)
            low_cost_payoff = np.matmul(low_cost_probabilities, np.matmul(
                self.matrix_A, np.transpose(high_cost_probabilities)))
            high_cost_payoff = np.matmul(low_cost_probabilities, np.matmul(
                self.matrix_B, np.transpose(high_cost_probabilities)))

            
            result = {"low_cost_probs": low_cost_probabilities,
                      "high_cost_probs": high_cost_probabilities,
                      "low_cost_payoff": low_cost_payoff,
                      "high_cost_payoff": high_cost_payoff}
            equilibria.append(result)
        return equilibria


class Strategy():
    """
    strategies can be static or they can come from neural nets. If NN, policy is nn.policy o.w. the static function
    """
    type = None
    env = None
    name = None
    nn = None
    state_adv_hist = None
    policy = None

    def __init__(self, strategy_type, model_or_func, name, first_price=132, state_adv_hist=None, action_step=None) -> None:
        """
        model_or_func: for static strategy is the function, for sb3 is the optimizer class
        """
        self.type = strategy_type
        self.name = name
        # self._env = environment
        self.state_adv_hist = gl.NUM_ADV_HISTORY if (
            state_adv_hist is None) else state_adv_hist
        
        self.action_step = action_step

        if strategy_type == StrategyType.neural_net:
            # self.nn = ModelOrFunc
            # self.policy = ModelOrFunc.policy
            # self.state_adv_hist = gl.NUM_ADV_HISTORY
            pass
        elif strategy_type == StrategyType.sb3_model:
            self.dir = f"{gl.MODELS_DIR}/{name}"
            self.model = model_or_func
            # self.policy = self.model.predict

        else:
            self.policy = model_or_func
            self.first_price = first_price

    def reset(self):
        pass

    def play(self, env, player=1):
        """
            Computes the price to be played in the environment, nn.step_action is the step size for pricing less than myopic
        """

        if self.type == StrategyType.neural_net:
            # state = self.env.get_state(
            #     self.env.stage, player, adv_hist=gl.num_adv_history)
            # normState = normalize_state(state=state)
            # probs = self.policy(normState)
            # distAction = Categorical(probs)
            # action = distAction.sample()
            # return compute_price(action=action.item(), action_step=gl.ACTION_STEP, demand=self.env.demandPotential[player][self.env.stage], cost=self.env.costs[player])
            pass
        elif self.type == StrategyType.sb3_model:
            if self.policy is None:
                self.policy= (self.model.load(self.dir,env=env)).predict
            state = env.get_state(
                stage=env.stage, player=player, adv_hist=self.state_adv_hist)
            action, _ = self.policy(state)
            #compute price for co model and disc model
            price= (env.myopic(player)-action[0]) if (self.action_step is None) else (env.myopic(player)-(self.action_step*action))

            if player==0:
                env.actions[env.stage]= (action[0] if(self.action_step is None) else (self.action_step*action))

            return price
        else:
            return self.policy(env, player, self.first_price)

    def play_against(self, env, adversary):
        """ 
        self is player 0 and adversary is layer 1. The environment should be specified. action_step for the neural netwroks should be set.
        output: tuple (payoff of low cost, payoff of high cost)
        """
        self.env = env

        state,_ = env.reset()
        while env.stage < (env.T):
            prices = [0, 0]
            prices[0], prices[1] = self.play(env, 0), adversary.play(env, 1)
            env.update_game_variables(prices)
            env.stage += 1

        return [sum(env.profit[0]), sum(env.profit[1])]

    def to_mixed_strategy(self):
        """
        Returns a MixedStrategy, Pr(self)=1
        """
        mix = MixedStrategy(probablities_lst=[1],
                            strategies_lst=[self])

        return mix


class MixedStrategy():
    strategies = []
    strategy_probs = None

    def __init__(self, strategies_lst, probablities_lst) -> None:
        self.strategies = strategies_lst
        self.strategy_probs = probablities_lst
        self.support_size=support_count(probablities_lst)

    def choose_strategy(self):
        if len(self.strategies) > 0:
            # adversaryDist = Categorical(torch.tensor(self._strategyProbs))
            # if not torch.is_tensor(self._strategyProbs):
            #     self._strategyProbs = torch.tensor(self._strategyProbs)
            # adversaryDist = Categorical(self._strategyProbs)
            # strategyInd = (adversaryDist.sample()).item()
            strategy_ind = np.random.choice(
                len(self.strategies), size=1, p=self.strategy_probs)
            return self.strategies[strategy_ind[0]]
        else:
            print("adversary's strategy can not be set!")
            return None
        
    def play_against(self, env, adversary):
        pass

    def __str__(self) -> str:
        s = ""
        for i in range(len(self.strategies)):
            if self.strategy_probs[i] > 0:
                s += f"{self.strategies[i].name}-{self.strategy_probs[i]:.2f},"
        return s


class StrategyType(Enum):
    static = 0
    neural_net = 1
    sb3_model = 2


def myopic(env, player, firstprice=0):
    """
        Adversary follows Myopic strategy
    """
    return env.myopic(player)


def const(env, player, firstprice):  # constant price strategy
    """
        Adversary follows Constant strategy
    """
    if env.stage == env.T-1:
        return env.myopic(player)
    return firstprice


def imit(env, player, firstprice):  # price imitator strategy
    if env.stage == 0:
        return firstprice
    if env.stage == env.T-1:
        return env.myopic(player)
    return env.prices[1-player][env.stage-1]


def fight(env, player, firstprice):  # simplified fighting strategy
    if env.stage == 0:
        return firstprice
    if env.stage == env.T-1:
        return env.myopic(player)
    # aspire = [ 207, 193 ] # aspiration level for demand potential
    aspire = [0, 0]
    for i in range(2):
        aspire[i] = (env.total_demand-env.costs[player] +
                     env.costs[1-player])/2

    D = env.demand_potential[player][env.stage]
    Asp = aspire[player]
    if D >= Asp:  # keep price; DANGER: price will never rise
        return env.prices[player][env.stage-1]
    # adjust to get to aspiration level using previous
    # opponent price; own price has to be reduced by twice
    # the negative amount D - Asp to getenv.demandPotential to Asp
    P = env.prices[1-player][env.stage-1] + 2*(D - Asp)
    # never price to high because even 125 gives good profits
    # P = min(P, 125)
    aspire_price = (env.total_demand+env.costs[0]+env.costs[1])/4
    P = min(P, int(0.95*aspire_price))

    return P


def fight_lb(env, player, firstprice):
    P = env.fight(player, firstprice)
    # never price less than production cost
    P = max(P, env.costs[player])
    return P

# sophisticated fighting strategy, compare fight()
# estimate *sales* of opponent as their target


def guess(env, player, firstprice):  # predictive fighting strategy
    if env.stage == 0:
        env.aspireDemand = [(env.total_demand/2 + env.costs[1]-env.costs[0]),
                            (env.total_demand/2 + env.costs[0]-env.costs[1])]  # aspiration level
        env.aspirePrice = (env.total_demand+env.costs[0]+env.costs[1])/4
        # first guess opponent sales as in monopoly ( sale= demand-price)
        env.saleGuess = [env.aspireDemand[0]-env.aspirePrice,
                         env.aspireDemand[1]-env.aspirePrice]

        return firstprice

    if env.stage == env.T-1:
        return env.myopic(player)

    D = env.demand_potential[player][env.stage]
    Asp = env.aspireDemand[player]

    if D >= Asp:  # keep price, but go slightly towards monopoly if good
        pmono = env.myopic(player)
        pcurrent = env.prices[player][env.stage-1]
        if pcurrent > pmono:  # shouldn't happen
            return pmono
        elif pcurrent > pmono-7:  # no change
            return pcurrent
        # current low price at 60%, be accommodating towards "collusion"
        return .6 * pcurrent + .4 * (pmono-7)

    # guess current *opponent price* from previous sales
    prevsales = env.demand_potential[1 -
                                     player][env.stage-1] - env.prices[1-player][env.stage-1]
    # adjust with weight alpha from previous guess
    alpha = .5
    newsalesguess = alpha * env.saleGuess[player] + (1-alpha)*prevsales
    # update
    env.saleGuess[player] = newsalesguess
    guessoppPrice = env.total_demand - D - newsalesguess
    P = guessoppPrice + 2*(D - Asp)

    if player == 0:
        P = min(P, 125)
    if player == 1:
        P = min(P, 130)
    return P


def monopolyPrice(demand, cost):  # myopic monopoly price
    """
        Computes Monopoly prices.
    """
    return (demand + cost) / 2
    # return (self.demandPotential[player][self.stage] + self.costs[player])/2

def prt(string):
    """
    writing the progres into a file instead of print
    """
    with open('progress.txt','a') as file:
        file.write("\n"+string)


def write_to_excel(file_name, new_row):
    """
    row includes:  name	ep	costs	adversary	agent_return	adv_return	agent_rewards	actions	agent_prices	adv_prices	agent_demands	adv_demands	lr	hist	total_stages	action_step	num_actions	gamma	stae_onehot	seed	num_procs	running_time
    """

    path = 'results.xlsx' if (file_name is None) else file_name
    wb = load_workbook(path)
    sheet = wb.active
    row = 2
    col = 1
    sheet.insert_rows(idx=row)

    for i in range(len(new_row)):
        sheet.cell(row=row, column=col+i).value = new_row[i]
    wb.save(path)

def write_results(new_row):
    write_to_excel('results.xlsx', new_row)
def write_agents(new_row):
    # name	ep	costs	adversary	expected_payoff	payoff_treshhold	lr	hist	total_stages	action_step	num_actions\
    #gamma	seed	num_procs	running_time	date
							
    write_to_excel('trained_agents.xlsx', new_row)


def support_count(list):
    """
    gets a list and returns the number of elements that are greater than zero
    """
    counter = 0
    for item in list:
        if item > 0:
            counter += 1
    return counter

def recover_probs(test):
    low_cost_probs, high_cost_probs, rest = test.split(")")
    low_cost_probs = low_cost_probs.split("(")[1]
    _, high_cost_probs = high_cost_probs.split("(")
    high_cost_probs = [float(Fraction(s)) for s in high_cost_probs.split(',')]
    low_cost_probs = [float(Fraction(s)) for s in low_cost_probs.split(',')]
    _, low_cost_support, high_cost_support = rest.split('[')
    high_cost_support, _ = high_cost_support.split(']')
    high_cost_support = [int(s) for s in high_cost_support.split(',')]
    low_cost_support, _ = low_cost_support.split(']')
    low_cost_support = [int(s) for s in low_cost_support.split(',')]
    return low_cost_probs, high_cost_probs, low_cost_support, high_cost_support


def return_distribution(number_players, cost_probs, cost_support):
    player_probabilities = [0] * number_players
    for index, support in enumerate(cost_support):
        player_probabilities[support] = cost_probs[support]
    return player_probabilities

def create_directories():
    if not os.path.exists(gl.MODELS_DIR):
        os.makedirs(gl.MODELS_DIR)
    if not os.path.exists(gl.LOG_DIR):
        os.makedirs(gl.LOG_DIR)
    if not os.path.exists(gl.MODELS_DIR):
        os.makedirs(gl.MODELS_DIR)
    if not os.path.exists("games"):
        os.makedirs("games")