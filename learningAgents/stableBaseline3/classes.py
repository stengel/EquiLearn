from enum import Enum
import numpy as np
import globals as gl
# import torch
# from torch.distributions import Categorical
from openpyxl import load_workbook


class Strategy():
    """
    strategies can be static or they can come from neural nets. If NN, policy is nn.policy o.w. the static function
    """
    type = None
    env = None
    name = None
    nn = None
    nn_hist = None
    policy = None

    def __init__(self, strategyType, NNorFunc, name, firstPrice=132) -> None:
        """
        Based on the type of strategy, the neuralnet or the Strategy Function  should be given as input. FirstPrice just applies to static strategies
        """
        self.type = strategyType
        self.name = name
        # self._env = environment

        if strategyType == StrategyType.neural_net:
            self.nn = NNorFunc
            self.policy = NNorFunc.policy
            self.nn_hist = gl.NUM_ADV_HISTORY
        else:
            self.policy = NNorFunc
            self.firstPrice = firstPrice

    def reset(self):
        pass

    def play(self, environment, player=1):
        """
            Computes the action to be played in the environment, nn.step_action is the step size for pricing less than myopic
        """
        self.env = environment
        if self.type == StrategyType.neural_net:
            # state = self.env.get_state(
            #     self.env.stage, player, adv_hist=gl.num_adv_history)
            # normState = normalize_state(state=state)
            # probs = self.policy(normState)
            # distAction = Categorical(probs)
            # action = distAction.sample()
            # return compute_price(action=action.item(), action_step=gl.ACTION_STEP, demand=self.env.demandPotential[player][self.env.stage], cost=self.env.costs[player])
            pass
        else:
            return self.policy(self.env, player, self.firstPrice)

    def play_against(self, env, adversary):
        """ 
        self is player 0 and adversary is layer 1. The environment should be specified. action_step for the neural netwroks should be set.
        output: tuple (payoff of low cost, payoff of high cost)
        """
        self.env = env

        state, reward, done = env.reset()
        while env.stage < (env.T):
            prices = [0, 0]
            prices[0], prices[1] = self.play(env, 0), adversary.play(env, 1)
            env.updatePricesProfitDemand(prices)
            env.stage += 1

        return [sum(env.profit[0]), sum(env.profit[1])]

    def to_mixed_strategy(self):
        """
        Returns a MixedStrategy, Pr(self)=1
        """
        mix = MixedStrategy(probablitiesArray=np.ones(1),
                            strategiesList=[self])

        return mix


class MixedStrategy():
    _strategies = []
    _strategyProbs = None

    def __init__(self, strategiesList=[], probablitiesArray=None) -> None:
        self._strategies = strategiesList
        self._strategyProbs = probablitiesArray

    def set_adversary_strategy(self):
        if len(self._strategies) > 0:
            # adversaryDist = Categorical(torch.tensor(self._strategyProbs))
            # if not torch.is_tensor(self._strategyProbs):
            #     self._strategyProbs = torch.tensor(self._strategyProbs)
            # adversaryDist = Categorical(self._strategyProbs)
            # strategyInd = (adversaryDist.sample()).item()
            strategyInd= np.random.choice(len(self._strategies), size=1, p= self._strategyProbs)
            return self._strategies[strategyInd[0]]
        else:
            print("adversary's strategy can not be set!")
            return None

    def __str__(self) -> str:
        s = ""
        for i in range(len(self._strategies)):
            if self._strategyProbs[i] > 0:
                s += f"{self._strategies[i].name}-{self._strategyProbs[i]:.2f},"
        return s


class StrategyType(Enum):
    static = 0
    neural_net = 1


def myopic(env, player, firstprice=0):
    """
        Adversary follows Myopic strategy
    """
    return env.myopic(player)


def const(env, player, firstprice):  # constant price strategy
    """
        Adversary follows Constant strategy
    """
    if env.stage == env.T-1:
        return env.myopic(player)
    return firstprice


def imit(env, player, firstprice):  # price imitator strategy
    if env.stage == 0:
        return firstprice
    if env.stage == env.T-1:
        return env.myopic(player)
    return env.prices[1-player][env.stage-1]


def fight(env, player, firstprice):  # simplified fighting strategy
    if env.stage == 0:
        return firstprice
    if env.stage == env.T-1:
        return env.myopic(player)
    # aspire = [ 207, 193 ] # aspiration level for demand potential
    aspire = [0, 0]
    for i in range(2):
        aspire[i] = (env.total_demand-env.costs[player] +
                     env.costs[1-player])/2

    D = env.demand_potential[player][env.stage]
    Asp = aspire[player]
    if D >= Asp:  # keep price; DANGER: price will never rise
        return env.prices[player][env.stage-1]
    # adjust to get to aspiration level using previous
    # opponent price; own price has to be reduced by twice
    # the negative amount D - Asp to getenv.demandPotential to Asp
    P = env.prices[1-player][env.stage-1] + 2*(D - Asp)
    # never price to high because even 125 gives good profits
    # P = min(P, 125)
    aspire_price = (env.total_demand+env.costs[0]+env.costs[1])/4
    P = min(P, int(0.95*aspire_price))

    return P


def fight_lb(env, player, firstprice):
    P = env.fight(player, firstprice)
    # never price less than production cost
    P = max(P, env.costs[player])
    return P

# sophisticated fighting strategy, compare fight()
# estimate *sales* of opponent as their target


def guess(env, player, firstprice):  # predictive fighting strategy
    if env.stage == 0:
        env.aspireDemand = [(env.totalDemand/2 + env.costs[1]-env.costs[0]),
                            (env.totalDemand/2 + env.costs[0]-env.costs[1])]  # aspiration level
        env.aspirePrice = (env.totalDemand+env.costs[0]+env.costs[1])/4
        # first guess opponent sales as in monopoly ( sale= demand-price)
        env.saleGuess = [env.aspireDemand[0]-env.aspirePrice,
                         env.aspireDemand[1]-env.aspirePrice]

        return firstprice

    if env.stage == env.T-1:
        return env.myopic(player)

    D = env.demand_potential[player][env.stage]
    Asp = env.aspireDemand[player]

    if D >= Asp:  # keep price, but go slightly towards monopoly if good
        pmono = env.myopic(player)
        pcurrent = env.prices[player][env.stage-1]
        if pcurrent > pmono:  # shouldn't happen
            return pmono
        elif pcurrent > pmono-7:  # no change
            return pcurrent
        # current low price at 60%, be accommodating towards "collusion"
        return .6 * pcurrent + .4 * (pmono-7)

    # guess current *opponent price* from previous sales
    prevsales = env.demand_potential[1 -
                                    player][env.stage-1] - env.prices[1-player][env.stage-1]
    # adjust with weight alpha from previous guess
    alpha = .5
    newsalesguess = alpha * env.saleGuess[player] + (1-alpha)*prevsales
    # update
    env.saleGuess[player] = newsalesguess
    guessoppPrice = env.total_demand - D - newsalesguess
    P = guessoppPrice + 2*(D - Asp)

    if player == 0:
        P = min(P, 125)
    if player == 1:
        P = min(P, 130)
    return P


def monopolyPrice(demand, cost):  # myopic monopoly price
    """
        Computes Monopoly prices.
    """
    return (demand + cost) / 2
    # return (self.demandPotential[player][self.stage] + self.costs[player])/2


def write_to_excel(new_row):
    """
    row includes:  name	ep	costs	adversary	agent_return	adv_return	agent_rewards	actions	agent_prices	adv_prices	agent_demands	adv_demands	lr	hist	total_stages	action_step	num_actions	gamma	stae_onehot	seed	num_procs	running_time
    """

    path = 'results.xlsx'
    wb = load_workbook(path)
    sheet = wb.active
    row = 2
    col = 1
    sheet.insert_rows(idx=row)

    for i in range(len(new_row)):
        sheet.cell(row=row, column=col+i).value = new_row[i]
    wb.save(path)